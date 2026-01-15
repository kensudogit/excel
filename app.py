"""
Excelファイル検索・抽出アプリケーション

このアプリケーションは、指定したフォルダ内のExcelファイル（.xlsx, .xls）から
複数のキーワードを検索し、検索結果を別のExcelブックに出力する機能を提供します。

主な機能:
- フォルダ内のExcelファイルからキーワード検索
- アップロードされたExcelファイルからキーワード検索
- 検索結果のExcelファイルへの出力（ハイパーリンク付き）
- セル詳細情報の取得
- Excelファイルの直接開き（Windows環境）
- 一括検索・置換機能（正規表現対応）

技術スタック:
- Flask: Webフレームワーク
- openpyxl: Excelファイルの読み書き
- pandas: データ処理（必要に応じて）
- flask-cors: CORS対応
"""
import os  # オペレーティングシステム関連の機能
import json  # JSONデータの処理
import re  # 正規表現処理
import shutil  # ファイル操作（コピーなど）
import subprocess  # 外部プロセスの実行
import platform  # プラットフォーム情報の取得
from pathlib import Path  # パス操作のためのクラス
from flask import Flask, request, jsonify, send_file  # Flask関連のインポート
from flask_cors import CORS  # CORS（Cross-Origin Resource Sharing）対応
import openpyxl  # Excelファイルの読み書きライブラリ
from openpyxl import Workbook  # Excelワークブックの作成
from openpyxl.styles import Font, PatternFill, Alignment  # Excelのスタイル設定
import pandas as pd  # データ分析ライブラリ（必要に応じて使用）
from datetime import datetime  # 日時処理

# ============================================================================
# オプションライブラリのインポート
# ============================================================================

# Hyperlinkクラスのインポート（オプション）
# Excelファイルにハイパーリンクを設定するために使用
# 利用できない場合は文字列でハイパーリンクを設定
try:
    from openpyxl.cell.hyperlink import Hyperlink
    HYPERLINK_AVAILABLE = True
except ImportError:
    try:
        # 別のインポート方法を試す（openpyxlのバージョンによって異なる可能性がある）
        from openpyxl.cell import Hyperlink
        HYPERLINK_AVAILABLE = True
    except ImportError:
        HYPERLINK_AVAILABLE = False
        # Vercel環境ではprint文が問題を引き起こす可能性があるため、ログ出力は後で行う
        # 警告は後でapp.loggerを使用して出力する

# Windows環境でExcelを操作するためのライブラリ（オプション）
# win32comを使用すると、Excelアプリケーションを直接操作できる
# 利用できない環境（Linux/Mac）でも動作するようにオプションとして扱う
try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False

# ============================================================================
# Flaskアプリケーションの初期化
# ============================================================================

# Flaskアプリケーションインスタンスの作成
app = Flask(__name__)

# CORS（Cross-Origin Resource Sharing）を有効化
# これにより、異なるドメインからのリクエストを許可する（完全公開モード）
CORS(app)

# Hyperlinkが利用できない場合の警告をログに出力
if not HYPERLINK_AVAILABLE:
    import logging
    logging.warning("Hyperlink class not available, will use string hyperlinks")

# ファイルアップロードサイズ制限を設定
# デフォルトは16MBだが、大きなExcelファイルに対応するため100MBに拡大
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB

# ============================================================================
# ディレクトリ設定
# ============================================================================

# アプリケーションのベースディレクトリを取得
# app.pyがあるディレクトリをベースディレクトリとして使用
BASE_DIR = Path(__file__).parent.resolve()

# アップロードファイルと結果ファイルの保存先を設定
# Vercel環境の場合は/tmpディレクトリを使用（Serverless Functionsの制約）
# ローカル環境の場合は、アプリケーションディレクトリ内に作成
if os.environ.get('VERCEL'):
    # Vercel環境: Serverless Functionsでは/tmpディレクトリのみ書き込み可能
    TMP_BASE = Path('/tmp')
    UPLOAD_FOLDER = TMP_BASE / 'uploads'  # アップロードされたファイルの一時保存先
    RESULTS_FOLDER = TMP_BASE / 'results'  # 検索結果のExcelファイルの保存先
else:
    # ローカル環境: アプリケーションディレクトリ内にフォルダを作成
    UPLOAD_FOLDER = BASE_DIR / 'uploads'  # アップロードされたファイルの保存先
    RESULTS_FOLDER = BASE_DIR / 'results'  # 検索結果のExcelファイルの保存先

# 環境変数で上書き可能（デプロイ環境などで設定）
# これにより、異なる環境で異なる保存先を指定できる
UPLOAD_FOLDER = Path(os.environ.get('UPLOAD_FOLDER', str(UPLOAD_FOLDER)))
RESULTS_FOLDER = Path(os.environ.get('RESULTS_FOLDER', str(RESULTS_FOLDER)))

# ディレクトリが存在しない場合は作成
# parents=True: 親ディレクトリも含めて作成
# exist_ok=True: 既に存在する場合はエラーを出さない
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
RESULTS_FOLDER.mkdir(parents=True, exist_ok=True)


def search_keywords_in_excel(file_path, keywords):
    """
    Excelファイル内でキーワードを検索する関数
    
    指定されたExcelファイルの全シートを走査し、各セルの値をチェックして
    キーワードが含まれているかどうかを判定します。
    大文字小文字を区別しない検索を行います。
    
    引数:
        file_path: 検索対象のExcelファイルのパス（Pathオブジェクトまたは文字列）
        keywords: 検索するキーワードのリスト（例: ['キーワード1', 'キーワード2']）
    
    戻り値:
        list: 検索結果のリスト。各要素は以下のキーを持つ辞書:
            - 'sheet': シート名
            - 'row': 行番号（1から始まる）
            - 'col': 列番号（1から始まる）
            - 'value': セルの値（文字列）
            - 'keyword': マッチしたキーワード
            - 'file': ファイルパス
    
    処理の流れ:
        1. Excelファイルを開く（data_only=Trueで計算式の結果を取得）
        2. 各シートを順に処理
        3. 各セルを走査し、値がNoneでない場合のみ処理
        4. セルの値を文字列に変換し、各キーワードと比較（大文字小文字を区別しない）
        5. マッチした場合は結果リストに追加
    """
    results = []
    try:
        # file_pathがPathオブジェクトの場合は文字列に変換
        # openpyxlは文字列形式のパスを期待するため
        file_path_str = str(file_path) if isinstance(file_path, Path) else file_path
        
        # Excelファイルを開く
        # data_only=True: 計算式の結果のみを取得（計算式自体は取得しない）
        wb = openpyxl.load_workbook(file_path_str, data_only=True)
        
        # 各シートを順に処理
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 各行を走査（values_only=Falseでセルオブジェクトを取得）
            # enumerate(..., start=1)で行番号を1から始める
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                # 各列（セル）を走査
                for col_idx, cell in enumerate(row, start=1):
                    # セルの値がNoneの場合はスキップ（空セル）
                    if cell.value is None:
                        continue
                    
                    # セルの値を文字列に変換
                    cell_value = str(cell.value)
                    
                    # 各キーワードをチェック
                    for keyword in keywords:
                        # 大文字小文字を区別しない検索
                        # セルの値とキーワードの両方を小文字に変換して比較
                        if keyword.lower() in cell_value.lower():
                            # マッチした場合は結果リストに追加
                            results.append({
                                'sheet': sheet_name,  # シート名
                                'row': row_idx,  # 行番号（1から始まる）
                                'col': col_idx,  # 列番号（1から始まる）
                                'value': cell_value,  # セルの値
                                'keyword': keyword,  # マッチしたキーワード
                                'file': file_path_str  # ファイルパス（後で上書きされる可能性がある）
                            })
        
        # Excelファイルを閉じる（メモリリークを防ぐ）
        wb.close()
    except Exception as e:
        # エラーが発生した場合は、エラー情報をログに記録
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error processing {file_path}: {error_trace}")
        app.logger.error(f"Error processing {file_path}: {error_trace}")
        # エラーが発生しても空のリストを返す（他のファイルの処理を継続）
    
    return results


def create_results_workbook(search_results, keywords):
    """
    検索結果をExcelブックに出力する関数
    
    検索結果のリストを受け取り、見やすい形式でExcelファイルに出力します。
    各検索結果に対して、ファイル名、シート名、行、列、セル値、キーワード、ファイルパスを
    記録します。また、ハイパーリンクを設定して、元のExcelファイルや該当セルに
    直接ジャンプできるようにします。
    
    引数:
        search_results: 検索結果のリスト（search_keywords_in_excelの戻り値）
        keywords: 検索に使用したキーワードのリスト（行の色分けに使用）
    
    戻り値:
        Workbook: 作成されたExcelワークブックオブジェクト
    
    処理の流れ:
        1. 新しいワークブックを作成
        2. ヘッダー行を設定（スタイル付き）
        3. 各検索結果を行として追加
        4. ハイパーリンクを設定（ファイル名、セル値、ファイルパス）
        5. キーワードに応じて行の色を変更
        6. 列幅を自動調整
    """
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active  # アクティブなワークシートを取得
    ws.title = "検索結果"  # シート名を設定
    
    # ヘッダー行の定義
    # 検索結果のExcelファイルに表示する列名
    headers = ['ファイル名', 'シート名', '行', '列', 'セル値', 'キーワード', 'ファイルパス']
    ws.append(headers)  # ヘッダー行を追加
    
    # ヘッダー行のスタイル設定
    # 背景色: 青色（#4472C4）、文字色: 白色、太字
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 各ヘッダーセルにスタイルを適用
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill  # 背景色を設定
        cell.font = header_font  # フォントを設定
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 中央揃え
    
    # ========================================================================
    # データ行の処理
    # ========================================================================
    # 各検索結果を行として追加し、ハイパーリンクを設定
    for result in search_results:
        file_path = result['file']  # 元のExcelファイルのパス
        file_path_obj = Path(file_path)
        
        # ファイルパスを絶対パスに変換（ハイパーリンク用）
        # ハイパーリンクは絶対パスでないと正しく動作しないため
        try:
            if file_path_obj.exists():
                # ファイルが存在する場合は、絶対パスに解決
                absolute_file_path = str(file_path_obj.resolve())
            else:
                # ファイルが存在しない場合（アップロードされたファイルなど）は、そのまま使用
                absolute_file_path = str(file_path_obj)
        except:
            # エラーが発生した場合は、元のパス文字列を使用
            absolute_file_path = str(file_path)
        
        # ハイパーリンク用のパス形式に変換
        # WindowsとLinux/Macで形式が異なるため、プラットフォームを判定
        if platform.system() == 'Windows':
            # Windowsの場合: file:///C:/path/to/file.xlsx の形式
            # バックスラッシュをスラッシュに変換
            hyperlink_path = absolute_file_path.replace('\\', '/')
            # file:///プレフィックスを追加（3つのスラッシュ）
            hyperlink_path = f"file:///{hyperlink_path}"
        else:
            # Linux/Macの場合: file:///path/to/file.xlsx の形式
            hyperlink_path = f"file://{absolute_file_path}"
        
        # 行データを構築
        # ファイル名、シート名、行、列、セル値、キーワード、ファイルパス
        row = [
            Path(result['file']).name,  # ファイル名のみ（パスから抽出）
            result['sheet'],  # シート名
            result['row'],  # 行番号
            result['col'],  # 列番号
            result['value'],  # セルの値
            result['keyword'],  # マッチしたキーワード
            result['file']  # ファイルパス（フルパス）
        ]
        ws.append(row)  # 行を追加
        
        # 現在の行番号を取得（ハイパーリンク設定用）
        current_row = ws.max_row
        
        # ====================================================================
        # ハイパーリンクの設定
        # ====================================================================
        
        # ファイル名のセル（1列目）にハイパーリンクを設定
        # クリックすると元のExcelファイルが開く
        file_name_cell = ws.cell(row=current_row, column=1)
        try:
            if HYPERLINK_AVAILABLE:
                # Hyperlinkオブジェクトを使用してハイパーリンクとツールチップを設定
                try:
                    file_name_cell.hyperlink = Hyperlink(
                        target=hyperlink_path,  # リンク先のパス
                        tooltip=f"クリックしてファイルを開く: {absolute_file_path}"  # ホバー時の説明
                    )
                except (TypeError, AttributeError):
                    # 古いバージョンのopenpyxlではtooltipパラメータがない場合がある
                    try:
                        file_name_cell.hyperlink = Hyperlink(target=hyperlink_path)
                    except Exception:
                        # Hyperlinkオブジェクトの作成に失敗した場合は文字列を直接設定
                        file_name_cell.hyperlink = hyperlink_path
            else:
                # Hyperlinkクラスが利用できない場合は文字列を直接設定
                file_name_cell.hyperlink = hyperlink_path
            
            # ハイパーリンクのスタイル設定（青色、下線付き）
            file_name_cell.font = Font(color="0563C1", underline="single")
        except Exception as e:
            # エラーが発生しても処理を継続（ログ出力のみ）
            # ハイパーリンクの設定に失敗しても、検索結果の出力は継続
            try:
                import logging
                logging.warning(f"Failed to set hyperlink for {file_path}: {str(e)}")
            except:
                pass
        
        # セル値のセル（5列目）に特定のセルへのハイパーリンクを設定
        # クリックすると元のExcelファイルが開き、該当セルに直接ジャンプする
        cell_value_cell = ws.cell(row=current_row, column=5)
        try:
            # ファイルパスが存在する場合、または絶対パスが取得できた場合はハイパーリンクを設定
            # アップロードされたファイルの場合、ファイル名のみの可能性があるが、可能な限りハイパーリンクを設定
            if absolute_file_path and (file_path_obj.exists() or os.path.isabs(absolute_file_path) or '\\' in absolute_file_path or '/' in absolute_file_path):
                # シート名とセル位置を含むハイパーリンクを作成
                sheet_name = result['sheet']  # シート名
                row_num = result['row']  # 行番号
                col_num = result['col']  # 列番号
                
                # 列番号をExcelの列文字（A, B, C...）に変換する内部関数
                # 例: 1 -> A, 2 -> B, 27 -> AA
                def number_to_excel_column(n):
                    """
                    数値をExcelの列文字に変換する関数
                    
                    引数:
                        n: 列番号（1から始まる）
                    
                    戻り値:
                        str: Excelの列文字（A, B, C, ..., Z, AA, AB, ...）
                    
                    アルゴリズム:
                        26進数のように変換（ただし、1から始まる）
                    """
                    result_col = ""
                    while n > 0:
                        n -= 1  # 0ベースに変換
                        result_col = chr(65 + (n % 26)) + result_col  # A=65, B=66, ...
                        n //= 26  # 次の桁へ
                    return result_col
                
                col_letter = number_to_excel_column(col_num)  # 列番号を列文字に変換
                
                # シート名に特殊文字が含まれている場合はシングルクォートで囲む
                # Excelのセル参照では、特殊文字を含むシート名はクォートで囲む必要がある
                if ' ' in sheet_name or '-' in sheet_name or any(c in sheet_name for c in ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')']):
                    cell_reference = f"'{sheet_name}'!{col_letter}{row_num}"  # クォート付き
                else:
                    cell_reference = f"{sheet_name}!{col_letter}{row_num}"  # クォートなし
                
                # セルへのジャンプを含むハイパーリンクパス
                # 形式: file:///path/to/file.xlsx#Sheet1!A1
                cell_hyperlink_path = f"{hyperlink_path}#{cell_reference}"
                
                if HYPERLINK_AVAILABLE:
                    try:
                        cell_value_cell.hyperlink = Hyperlink(
                            target=cell_hyperlink_path,
                            tooltip=f"クリックしてセル {cell_reference} にジャンプ: {absolute_file_path}"
                        )
                    except (TypeError, AttributeError):
                        try:
                            cell_value_cell.hyperlink = Hyperlink(target=cell_hyperlink_path)
                        except Exception:
                            cell_value_cell.hyperlink = cell_hyperlink_path
                else:
                    cell_value_cell.hyperlink = cell_hyperlink_path
                
                # ハイパーリンクのスタイル設定（青色、下線付き）
                cell_value_cell.font = Font(color="0563C1", underline="single")
            else:
                # ファイルパスが取得できない場合でも、少なくともフォントを設定
                # （アップロードされたファイルの場合など）
                cell_value_cell.font = Font(color="000000")  # 通常の黒色
        except Exception as e:
            # エラーが発生しても処理を継続（ログ出力のみ）
            try:
                import logging
                logging.warning(f"Failed to set cell value hyperlink: {str(e)}")
            except:
                pass
        
        # ファイルパスのセル（7列目）にもハイパーリンクを設定
        # クリックすると元のExcelファイルが開く
        file_path_cell = ws.cell(row=current_row, column=7)
        try:
            if HYPERLINK_AVAILABLE:
                # Hyperlinkオブジェクトを使用してハイパーリンクとツールチップを設定
                try:
                    file_path_cell.hyperlink = Hyperlink(
                        target=hyperlink_path,
                        tooltip=f"クリックしてファイルを開く: {absolute_file_path}"
                    )
                except (TypeError, AttributeError):
                    # 古いバージョンのopenpyxlではtooltipパラメータがない場合がある
                    try:
                        file_path_cell.hyperlink = Hyperlink(target=hyperlink_path)
                    except Exception:
                        # Hyperlinkオブジェクトの作成に失敗した場合は文字列を直接設定
                        file_path_cell.hyperlink = hyperlink_path
            else:
                # Hyperlinkクラスが利用できない場合は文字列を直接設定
                file_path_cell.hyperlink = hyperlink_path
            
            # ハイパーリンクのスタイル設定（青色、下線付き）
            file_path_cell.font = Font(color="0563C1", underline="single")
        except Exception as e:
            # エラーが発生しても処理を継続（ログ出力のみ）
            try:
                import logging
                logging.warning(f"Failed to set hyperlink for file path {file_path}: {str(e)}")
            except:
                pass
        
        # ====================================================================
        # 行の色分け（キーワードに応じて）
        # ====================================================================
        # 各キーワードに異なる色を割り当てて、検索結果を見やすくする
        keyword_colors = {
            keywords[0]: "FFE6E6" if len(keywords) > 0 else "FFFFFF",  # 1番目のキーワード: 薄い赤
            keywords[1]: "E6F3FF" if len(keywords) > 1 else "FFFFFF",  # 2番目のキーワード: 薄い青
            keywords[2]: "E6FFE6" if len(keywords) > 2 else "FFFFFF",  # 3番目のキーワード: 薄い緑
        }
        # マッチしたキーワードに対応する色を取得（デフォルトは白）
        fill_color = keyword_colors.get(result['keyword'], "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # 行の各セルに背景色を設定
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=current_row, column=col)
            # ハイパーリンクが設定されているセル（1列目: ファイル名、5列目: セル値、7列目: ファイルパス）の
            # フォント色は保持（背景色のみ設定）
            if col not in [1, 5, 7] or not cell.hyperlink:
                cell.fill = fill
    
    # ========================================================================
    # 列幅の自動調整
    # ========================================================================
    # 各列の内容に応じて列幅を自動調整し、見やすくする
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # 列文字（A, B, C, ...）を取得
        
        # 列内の各セルの最大文字数を取得
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass  # エラーが発生した場合はスキップ
        
        # 列幅を調整（最大50文字まで、最小2文字の余白を追加）
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    return wb


def normalize_path(path_str):
    """
    パス文字列を正規化する関数
    
    異なる形式で入力されたパス（相対パス、絶対パス、Windows形式、Unix形式など）を
    統一された形式に変換します。これにより、異なる環境や入力方法からの
    パスを正しく処理できるようになります。
    
    引数:
        path_str: 正規化するパス文字列
    
    戻り値:
        Path: 正規化されたPathオブジェクト。無効な場合はNone
    
    処理内容:
        1. 前後の空白を削除
        2. Pathオブジェクトに変換
        3. 存在する場合は絶対パスに解決
        4. 存在しない場合でも可能な限り絶対パスに変換を試みる
    
    使用例:
        normalize_path("C:\\Users\\Documents") -> Path("C:/Users/Documents")
        normalize_path("../folder") -> Path("/absolute/path/to/folder")
    """
    if not path_str:
        return None
    
    # 文字列の前後の空白を削除
    path_str = path_str.strip()
    
    # 空文字列の場合はNoneを返す
    if not path_str:
        return None
    
    # パスを正規化（Pathオブジェクトを使用）
    try:
        # まずPathオブジェクトに変換
        # Pathオブジェクトは、WindowsとUnixの両方のパス形式を自動的に処理
        path_obj = Path(path_str)
        
        # 絶対パスに変換（存在する場合）
        if path_obj.exists():
            # ファイル/フォルダが存在する場合は、絶対パスに解決
            path_obj = path_obj.resolve()
        else:
            # 存在しない場合でも、絶対パスに変換を試みる
            # これにより、相対パスを絶対パスに変換できる
            try:
                path_obj = path_obj.resolve()
            except (OSError, RuntimeError):
                # 解決できない場合は、元のパスを使用
                # 例: 存在しないパスや、無効なパス形式の場合
                pass
        
        return path_obj
    except Exception as e:
        # エラーが発生した場合は警告をログに記録
        app.logger.warning(f"Path normalization error for '{path_str}': {str(e)}")
        # エラーが発生した場合でも、元のパスをPathオブジェクトとして返す
        # これにより、処理を継続できる
        return Path(path_str)


# ============================================================================
# APIエンドポイント
# ============================================================================

@app.route('/api/search', methods=['POST'])
def search_excel_files():
    """
    指定フォルダ内のExcelファイルを検索するAPIエンドポイント
    
    リクエスト:
        POST /api/search
        Content-Type: application/json
        Body: {
            "folder_path": "C:\\Users\\Documents\\ExcelFiles",
            "keywords": ["キーワード1", "キーワード2", ...]
        }
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "results": [
                {
                    "file": "ファイルパス",
                    "sheet": "シート名",
                    "row": 行番号,
                    "col": 列番号,
                    "value": "セルの値",
                    "keyword": "マッチしたキーワード"
                },
                ...
            ],
            "total_matches": マッチ数,
            "files_searched": 検索したファイル数,
            "output_file": "結果ファイル名"
        }
        
        エラー時 (400, 404, 500):
        {
            "success": false,
            "error": "エラーメッセージ",
            "suggestion": "解決方法の提案"
        }
    
    処理の流れ:
        1. リクエストデータの検証
        2. フォルダパスの正規化と存在確認
        3. フォルダ内のExcelファイルを検索
        4. 各ファイルに対してキーワード検索を実行
        5. 検索結果をExcelファイルに出力
        6. 結果をJSON形式で返す
    """
    try:
        # リクエストデータの取得
        if not request.is_json:
            return jsonify({'success': False, 'error': 'リクエストはJSON形式である必要があります'}), 400
        
        data = request.get_json()
        if data is None:
            return jsonify({'success': False, 'error': 'リクエストデータが空です'}), 400
        
        folder_path = data.get('folder_path', '')
        keywords = data.get('keywords', [])
        
        if not folder_path:
            return jsonify({'success': False, 'error': 'フォルダパスが指定されていません'}), 400
        
        if not keywords or len(keywords) == 0:
            return jsonify({'success': False, 'error': 'キーワードが指定されていません'}), 400
        
        # パスを正規化
        original_path = folder_path
        folder = normalize_path(folder_path)
        
        # ログ出力（デバッグ用）
        app.logger.info(f"Search request - Original path: '{original_path}'")
        app.logger.info(f"Search request - Normalized path: '{folder}'")
        app.logger.info(f"Search request - Path exists: {folder.exists() if folder else False}")
        app.logger.info(f"Search request - Current working directory: {os.getcwd()}")
        
        if not folder:
            return jsonify({
                'success': False, 
                'error': f'フォルダパスが無効です: {original_path}',
                'suggestion': '絶対パス（例: C:\\Users\\Documents\\ExcelFiles）を入力してください'
            }), 400
        
        if not folder.exists():
            # より詳細なエラーメッセージを提供
            error_msg = f'指定されたフォルダが見つかりません: {original_path}'
            suggestion = '以下の点を確認してください:\n'
            suggestion += '1. フォルダパスが正しいか確認してください\n'
            suggestion += '2. 絶対パス（例: C:\\Users\\Documents\\ExcelFiles）を使用してください\n'
            suggestion += '3. 別のPCにデプロイした場合は、そのPC上に存在するフォルダパスを入力してください\n'
            suggestion += '4. フォルダ選択ボタンを使用して、正しいフォルダを選択してください'
            
            app.logger.error(f"Folder not found: '{original_path}' (normalized: '{folder}')")
            return jsonify({
                'success': False, 
                'error': error_msg,
                'suggestion': suggestion,
                'original_path': original_path,
                'normalized_path': str(folder)
            }), 404
        
        if not folder.is_dir():
            return jsonify({
                'success': False, 
                'error': f'指定されたパスはフォルダではありません: {original_path}',
                'suggestion': 'フォルダを指定してください（ファイルではなく）'
            }), 400
        
        # Excelファイルを検索
        excel_files = list(folder.glob('*.xlsx')) + list(folder.glob('*.xls'))
        
        app.logger.info(f"Found {len(excel_files)} Excel files in folder: {folder}")
        
        if not excel_files:
            # フォルダ内のファイル一覧を取得（デバッグ用）
            all_files = list(folder.iterdir())
            file_list = [f.name for f in all_files if f.is_file()][:10]  # 最初の10個のみ
            
            error_msg = f'フォルダ内にExcelファイルが見つかりませんでした: {original_path}'
            suggestion = '以下の点を確認してください:\n'
            suggestion += '1. フォルダ内に.xlsxまたは.xlsファイルが存在するか確認してください\n'
            suggestion += '2. ファイル名にスペースや特殊文字が含まれていないか確認してください'
            
            if file_list:
                suggestion += f'\nフォルダ内のファイル（最初の10個）: {", ".join(file_list)}'
            
            app.logger.warning(f"No Excel files found in folder: '{folder}' (files in folder: {len(all_files)})")
            return jsonify({
                'success': False, 
                'error': error_msg,
                'suggestion': suggestion,
                'folder_path': str(folder),
                'files_in_folder': file_list
            }), 404
        
        # 各ファイルを検索
        all_results = []
        for excel_file in excel_files:
            try:
                results = search_keywords_in_excel(excel_file, keywords)
                all_results.extend(results)
            except Exception as e:
                print(f"Error processing {excel_file}: {str(e)}")
                continue
        
        # 結果をExcelブックに出力
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = RESULTS_FOLDER / f'search_results_{timestamp}.xlsx'
            wb = create_results_workbook(all_results, keywords)
            wb.save(output_file)
        except Exception as e:
            print(f"Error creating workbook: {str(e)}")
            # ブック作成に失敗しても検索結果は返す
        
        # 結果をJSON形式で返す
        # 相対パスとして返す（RESULTS_FOLDERからの相対パス）
        if 'output_file' in locals() and output_file:
            # RESULTS_FOLDERからの相対パスを取得
            try:
                output_file_str = str(output_file.relative_to(RESULTS_FOLDER))
            except ValueError:
                # 相対パスにできない場合は、ファイル名のみを返す
                output_file_str = output_file.name
        else:
            output_file_str = None
        
        return jsonify({
            'success': True,
            'results': all_results,
            'total_matches': len(all_results),
            'files_searched': len(excel_files),
            'output_file': output_file_str
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in search_excel_files: {error_trace}")
        app.logger.error(f"Error in search_excel_files: {error_trace}")
        response = jsonify({
            'success': False,
            'error': f'検索中にエラーが発生しました: {str(e)}'
        })
        response.headers['Content-Type'] = 'application/json'
        return response, 500


@app.route('/api/search-files', methods=['POST'])
def search_excel_files_upload():
    """
    アップロードされたExcelファイルを検索するAPIエンドポイント
    
    このエンドポイントは、ブラウザから直接アップロードされたExcelファイルを
    検索するために使用されます。フォルダパスではなく、ファイル自体を
    アップロードして検索する場合に使用します。
    
    リクエスト:
        POST /api/search-files
        Content-Type: multipart/form-data
        Form Data:
            - keywords: JSON文字列（キーワードの配列）
            - files: Excelファイル（複数可）
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "results": [...],
            "total_matches": マッチ数,
            "files_searched": 検索したファイル数,
            "output_file": "結果ファイル名"
        }
    
    処理の流れ:
        1. アップロードされたファイルを取得
        2. Excelファイルのみをフィルタリング
        3. 一時ファイルとして保存
        4. 各ファイルに対してキーワード検索を実行
        5. 検索結果をExcelファイルに出力
        6. 一時ファイルを削除
        7. 結果をJSON形式で返す
    
    注意:
        - アップロードされたファイルは一時的に保存され、処理後に削除されます
        - ファイル名のみが結果に記録されます（元のパスは取得できません）
    """
    try:
        app.logger.info(f"Received request to /api/search-files")
        app.logger.info(f"Request method: {request.method}")
        app.logger.info(f"Request content type: {request.content_type}")
        app.logger.info(f"Request form keys: {list(request.form.keys())}")
        app.logger.info(f"Request files keys: {list(request.files.keys())}")
        
        # リクエストサイズのチェック（Vercelの制限を考慮）
        # VercelのServerless Functionsでは、リクエストサイズに制限がある
        content_length = request.content_length
        if content_length and content_length > 100 * 1024 * 1024:  # 100MB
            return jsonify({
                'success': False,
                'error': 'ファイルサイズが大きすぎます。100MB以下にしてください。'
            }), 413
        
        # キーワードの取得
        keywords_json = request.form.get('keywords', '[]')
        if not keywords_json:
            keywords_json = '[]'
        
        try:
            keywords = json.loads(keywords_json)
        except json.JSONDecodeError as e:
            app.logger.error(f"JSON decode error: {str(e)}, keywords_json: {keywords_json}")
            return jsonify({
                'success': False,
                'error': f'キーワードの形式が正しくありません: {str(e)}'
            }), 400
        
        if not keywords or len(keywords) == 0:
            return jsonify({'success': False, 'error': 'キーワードが指定されていません'}), 400
        
        # アップロードされたファイルの取得
        # 'files'キーが存在しない場合や、ファイルが空の場合はエラー
        if 'files' not in request.files:
            # ファイルがアップロードされていない場合
            app.logger.warning("No 'files' key in request.files")
            return jsonify({
                'success': False,
                'error': 'ファイルが指定されていません。Excelファイルをアップロードしてください。'
            }), 400
        
        uploaded_files = request.files.getlist('files')
        if not uploaded_files:
            app.logger.warning("uploaded_files is None or empty")
            return jsonify({
                'success': False,
                'error': 'ファイルが指定されていません。Excelファイルをアップロードしてください。'
            }), 400
        
        # 空のファイル名を除外
        uploaded_files = [f for f in uploaded_files if f.filename and f.filename.strip()]
        
        if len(uploaded_files) == 0:
            app.logger.warning("No valid files in uploaded_files")
            return jsonify({
                'success': False,
                'error': '有効なファイルがアップロードされていません。Excelファイル（.xlsx, .xls）をアップロードしてください。'
            }), 400
        
        # Excelファイルのみをフィルタリング
        excel_files = []
        for file in uploaded_files:
            if file.filename == '':
                continue
            if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
                excel_files.append(file)
        
        if not excel_files:
            return jsonify({'success': False, 'error': 'Excelファイルが見つかりませんでした'}), 404
        
        # 各ファイルを一時保存して検索
        all_results = []
        temp_file_paths = []
        
        for excel_file in excel_files:
            try:
                # 一時ファイルとして保存（ファイル名の重複を防ぐため、タイムスタンプを追加）
                import time
                timestamp = int(time.time() * 1000)
                safe_filename = f"{timestamp}_{excel_file.filename}"
                temp_file = UPLOAD_FOLDER / safe_filename
                excel_file.save(str(temp_file))
                temp_file_paths.append(temp_file)
                
                # 検索実行
                results = search_keywords_in_excel(temp_file, keywords)
                # ファイル名を元のファイル名に設定（パスではなくファイル名のみ）
                original_filename = excel_file.filename
                for result in results:
                    result['file'] = original_filename
                all_results.extend(results)
            except Exception as e:
                import traceback
                error_trace = traceback.format_exc()
                print(f"Error processing {excel_file.filename}: {error_trace}")
                app.logger.error(f"Error processing {excel_file.filename}: {error_trace}")
                continue
        
        # 結果をExcelブックに出力
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = RESULTS_FOLDER / f'search_results_{timestamp}.xlsx'
            wb = create_results_workbook(all_results, keywords)
            wb.save(output_file)
        except Exception as e:
            print(f"Error creating workbook: {str(e)}")
        
        # 一時ファイルを削除
        for temp_file in temp_file_paths:
            try:
                if temp_file.exists():
                    temp_file.unlink()
            except Exception as e:
                print(f"Error deleting temp file {temp_file}: {str(e)}")
        
        # 結果をJSON形式で返す
        # 相対パスとして返す（RESULTS_FOLDERからの相対パス）
        if 'output_file' in locals() and output_file:
            # RESULTS_FOLDERからの相対パスを取得
            try:
                output_file_str = str(output_file.relative_to(RESULTS_FOLDER))
            except ValueError:
                # 相対パスにできない場合は、ファイル名のみを返す
                output_file_str = output_file.name
        else:
            output_file_str = None
        
        app.logger.info(f"Search completed: {len(all_results)} matches found in {len(excel_files)} files")
        return jsonify({
            'success': True,
            'results': all_results,
            'total_matches': len(all_results),
            'files_searched': len(excel_files),
            'output_file': output_file_str
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error in search_excel_files_upload: {error_trace}")
        app.logger.error(f"Error in search_excel_files_upload: {error_trace}")
        response = jsonify({
            'success': False,
            'error': f'検索中にエラーが発生しました: {str(e)}',
            'traceback': error_trace if app.debug else None
        })
        response.headers['Content-Type'] = 'application/json'
        return response, 500


@app.route('/api/get-cell-details', methods=['POST'])
def get_cell_details():
    """
    特定のセルの詳細情報を取得するAPIエンドポイント
    
    検索結果で見つかったセルの周辺情報（前後の行など）を取得します。
    これにより、セルの文脈を理解しやすくなります。
    
    リクエスト:
        POST /api/get-cell-details
        Content-Type: application/json
        Body: {
            "file_path": "ファイルパス",
            "sheet_name": "シート名",
            "row": 行番号,
            "col": 列番号,
            "keyword": "キーワード",
            "context_rows": 5  // 前後何行表示するか（オプション、デフォルト: 5）
        }
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "file_name": "ファイル名",
            "sheet_name": "シート名",
            "target_cell": {
                "row": 行番号,
                "col": 列番号,
                "value": "セルの値",
                "keyword": "キーワード"
            },
            "context": [
                [
                    {
                        "row": 行番号,
                        "col": 列番号,
                        "value": "セルの値",
                        "is_target": true/false,
                        "is_header": true/false
                    },
                    ...
                ],
                ...
            ],
            "max_row": シートの最大行数,
            "max_col": シートの最大列数
        }
    
    処理の流れ:
        1. ファイルの存在確認
        2. Excelファイルを開く
        3. シートの存在確認
        4. 対象セルと周辺セルの情報を取得
        5. JSON形式で返す
    """
    try:
        data = request.json
        file_path = data.get('file_path', '')
        sheet_name = data.get('sheet_name', '')
        row = data.get('row', 0)
        col = data.get('col', 0)
        keyword = data.get('keyword', '')
        context_rows = data.get('context_rows', 5)  # 前後何行表示するか
        
        if not file_path or not sheet_name or not row or not col:
            return jsonify({'success': False, 'error': '必要なパラメータが不足しています'}), 400
        
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return jsonify({'success': False, 'error': 'ファイルが見つかりません'}), 404
        
        wb = openpyxl.load_workbook(file_path_obj, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'error': 'シートが見つかりません'}), 404
        
        sheet = wb[sheet_name]
        
        # 周辺のセル情報を取得
        context_data = []
        start_row = max(1, row - context_rows)
        end_row = min(sheet.max_row, row + context_rows)
        
        for r in range(start_row, end_row + 1):
            row_data = []
            for c in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r, column=c)
                cell_info = {
                    'row': r,
                    'col': c,
                    'value': str(cell.value) if cell.value is not None else '',
                    'is_target': (r == row and c == col),
                    'is_header': (r == 1)
                }
                row_data.append(cell_info)
            context_data.append(row_data)
        
        # ヒットしたセルの詳細情報
        target_cell = sheet.cell(row=row, column=col)
        
        result = {
            'success': True,
            'file_name': file_path_obj.name,
            'sheet_name': sheet_name,
            'target_cell': {
                'row': row,
                'col': col,
                'value': str(target_cell.value) if target_cell.value is not None else '',
                'keyword': keyword
            },
            'context': context_data,
            'max_row': sheet.max_row,
            'max_col': sheet.max_column
        }
        
        wb.close()
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/download-results', methods=['GET'])
def download_results():
    """
    検索結果のExcelファイルをダウンロードするAPIエンドポイント
    
    検索実行時に生成された結果Excelファイルをダウンロードします。
    ファイルパスは相対パス（ファイル名のみ）または絶対パスで指定できます。
    
    リクエスト:
        GET /api/download-results?file_path=ファイル名
    
    レスポンス:
        成功時 (200):
        - Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        - ファイルのダウンロード
    
        エラー時 (400, 404, 500):
        {
            "success": false,
            "error": "エラーメッセージ",
            "available_files": ["利用可能なファイル一覧"]
        }
    
    処理の流れ:
        1. ファイルパスを取得（URLデコード）
        2. パスを正規化
        3. RESULTS_FOLDERからの相対パスとして処理
        4. ファイルの存在確認（見つからない場合はファイル名で検索）
        5. ファイルをダウンロードとして送信
    
    注意:
        - ファイルが見つからない場合、ファイル名で部分一致検索も試行します
        - 最新のファイル（タイムスタンプが新しいもの）を優先的に選択します
    """
    try:
        from urllib.parse import unquote
        
        file_path = request.args.get('file_path', '')
        if not file_path:
            return jsonify({'success': False, 'error': 'ファイルパスが指定されていません'}), 400
        
        # URLデコード
        file_path = unquote(file_path)
        
        app.logger.info(f"Download request - Original file_path: {file_path}")
        
        # パスの正規化
        # バックスラッシュをスラッシュに変換
        file_path_normalized = file_path.replace('\\', '/')
        
        # RESULTS_FOLDERを絶対パスに変換
        results_folder_abs = RESULTS_FOLDER.resolve()
        
        # 相対パスの場合は、resultsフォルダからのパスとして処理
        if not os.path.isabs(file_path_normalized):
            # 'results/' または 'results\' プレフィックスを削除（既に含まれている場合）
            # 大文字小文字を区別せずに削除
            file_path_lower = file_path_normalized.lower()
            if file_path_lower.startswith('results/'):
                file_path_normalized = file_path_normalized[len('results/'):]
            elif file_path_lower.startswith('results\\'):
                file_path_normalized = file_path_normalized[len('results\\'):]
            
            # 先頭のスラッシュやバックスラッシュを削除
            file_path_normalized = file_path_normalized.lstrip('/\\')
            
            # 空の場合はエラー
            if not file_path_normalized:
                return jsonify({'success': False, 'error': 'ファイルパスが無効です'}), 400
            
            # RESULTS_FOLDER（絶対パス）からの相対パスとして処理
            file_path_obj = results_folder_abs / file_path_normalized
        else:
            file_path_obj = Path(file_path_normalized).resolve()
        
        # ファイルパスを絶対パスに変換（確実に）
        file_path_obj = file_path_obj.resolve()
        
        app.logger.info(f"Download request - Normalized path: {file_path_normalized}")
        app.logger.info(f"Download request - Full path (abs): {file_path_obj}")
        app.logger.info(f"Download request - RESULTS_FOLDER (abs): {results_folder_abs}")
        app.logger.info(f"Download request - File exists: {file_path_obj.exists()}")
        
        # ファイルが存在するか確認
        if not file_path_obj.exists():
            # RESULTS_FOLDER内の全ファイルをリストアップ（デバッグ用）
            available_files = []
            if results_folder_abs.exists():
                available_files = list(results_folder_abs.glob('*.xlsx'))
                app.logger.error(f"Available files in RESULTS_FOLDER: {[f.name for f in available_files]}")
                app.logger.error(f"Looking for: {file_path_normalized}")
                app.logger.error(f"Full path attempted: {file_path_obj}")
            
            # ファイル名が一致するファイルを探す（フォールバック）
            # ファイル名のみで比較（パス情報を無視）
            file_name_only = Path(file_path_normalized).name
            matching_files = [f for f in available_files if f.name == file_name_only]
            if matching_files:
                file_path_obj = matching_files[0]
                app.logger.info(f"Found file by name match: {file_path_obj}")
            else:
                # さらに、ファイル名の部分一致も試す
                matching_files = [f for f in available_files if file_name_only in f.name or f.name in file_name_only]
                if matching_files:
                    # 最新のファイルを選択（タイムスタンプが新しいもの）
                    matching_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                    file_path_obj = matching_files[0]
                    app.logger.info(f"Found file by partial name match: {file_path_obj}")
                else:
                    return jsonify({
                        'success': False, 
                        'error': f'ファイルが見つかりません: {file_path_obj}',
                        'requested_path': file_path,
                        'normalized_path': file_path_normalized,
                        'file_name_only': file_name_only,
                        'results_folder': str(results_folder_abs),
                        'available_files': [f.name for f in available_files]
                    }), 404
        
        return send_file(
            str(file_path_obj),
            as_attachment=True,
            download_name=file_path_obj.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error in download_results: {error_trace}")
        return jsonify({'success': False, 'error': f'ダウンロード中にエラーが発生しました: {str(e)}'}), 500


@app.route('/api/open-excel-file', methods=['POST'])
def open_excel_file():
    """
    Excelファイルを開くAPIエンドポイント（Windows環境）
    
    検索結果から元のExcelファイルを開き、該当するシートとセルに
    直接ジャンプします。Windows環境では、win32comを使用して
    Excelアプリケーションを直接操作できます。
    
    リクエスト:
        POST /api/open-excel-file
        Content-Type: application/json
        Body: {
            "file_path": "ファイルパス",
            "sheet_name": "シート名",
            "row": 行番号,
            "col": 列番号
        }
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "message": "Excelファイルを開きました（シートとセルに移動しました）"
        }
    
    処理の流れ:
        1. ファイルの存在確認
        2. Windows環境の場合:
           - win32comが利用可能: Excelアプリケーションを起動し、シートとセルに移動
           - win32comが利用不可: 通常の方法でファイルを開く
        3. その他の環境: プラットフォーム固有のコマンドでファイルを開く
    
    注意:
        - Windows環境以外では、シートとセルへのジャンプはできません
        - win32comが利用できない場合は、ファイルを開くだけです
    """
    try:
        data = request.json
        file_path = data.get('file_path', '')
        sheet_name = data.get('sheet_name', '')
        row = data.get('row', 0)
        col = data.get('col', 0)
        
        if not file_path:
            return jsonify({'success': False, 'error': 'ファイルパスが指定されていません'}), 400
        
        file_path_obj = Path(file_path)
        if not file_path_obj.exists():
            return jsonify({'success': False, 'error': 'ファイルが見つかりません'}), 404
        
        # Windows環境でExcelファイルを開く
        if platform.system() == 'Windows':
            # 特定のシートとセルに移動する場合は、COM経由でExcelを操作
            if WIN32COM_AVAILABLE and sheet_name and row > 0 and col > 0:
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = True
                    wb = excel.Workbooks.Open(str(file_path_obj))
                    
                    if sheet_name:
                        ws = wb.Worksheets(sheet_name)
                        ws.Activate()
                        if row > 0 and col > 0:
                            ws.Cells(row, col).Select()
                    
                    return jsonify({
                        'success': True,
                        'message': 'Excelファイルを開きました（シートとセルに移動しました）'
                    })
                except Exception as e:
                    # COM操作に失敗した場合は、通常の方法でファイルを開く
                    print(f"COM操作に失敗: {str(e)}")
                    os.startfile(str(file_path_obj))
                    return jsonify({
                        'success': True,
                        'message': 'Excelファイルを開きました'
                    })
            else:
                # 通常の方法でファイルを開く
                os.startfile(str(file_path_obj))
                return jsonify({
                    'success': True,
                    'message': 'Excelファイルを開きました'
                })
        else:
            # Windows以外の環境
            if platform.system() == 'Darwin':  # macOS
                subprocess.Popen(['open', str(file_path_obj)])
            else:  # Linux
                subprocess.Popen(['xdg-open', str(file_path_obj)])
            return jsonify({
                'success': True,
                'message': 'Excelファイルを開きました'
            })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/search-replace', methods=['POST'])
def search_replace_files():
    """
    フォルダ内の複数ファイルに対して一括検索・置換を実行するAPIエンドポイント
    
    指定したフォルダ内の複数のファイル（テキストファイル、Excelファイル）に対して
    検索・置換を一括実行します。正規表現にも対応しています。
    
    リクエスト:
        POST /api/search-replace
        Content-Type: application/json
        Body: {
            "folder_path": "フォルダパス",
            "search_pattern": "検索パターン",
            "replace_pattern": "置換パターン",
            "use_regex": true/false,  // 正規表現を使用するか
            "file_extensions": [".txt", ".csv", ...],  // 対象ファイル拡張子
            "preview_only": true/false  // プレビューのみか、実際に置換するか
        }
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "results": [
                {
                    "file_path": "ファイルパス",
                    "file_name": "ファイル名",
                    "matches": [
                        {
                            "line": 行番号,
                            "start": 開始位置,
                            "end": 終了位置,
                            "match_text": "マッチしたテキスト",
                            "line_content": "行の内容",
                            "context_before": "前のコンテキスト",
                            "context_after": "後のコンテキスト",
                            "sheet": "シート名（Excelの場合）",
                            "column": "列文字（Excelの場合）"
                        },
                        ...
                    ],
                    "total_matches": マッチ数,
                    "replaced": true/false,
                    "backup_path": "バックアップファイルパス"
                },
                ...
            ],
            "total_files": 対象ファイル数,
            "files_with_matches": マッチしたファイル数,
            "total_replacements": 置換数,
            "preview_only": true/false
        }
    
    処理の流れ:
        1. フォルダの存在確認
        2. 対象ファイルを取得（指定された拡張子のファイル）
        3. 正規表現パターンのコンパイル
        4. 各ファイルを処理:
           - Excelファイル: openpyxlを使用して処理
           - テキストファイル: 通常のファイル操作で処理
        5. プレビューモードでない場合:
           - バックアップを作成
           - 置換を実行
           - ファイルを保存
        6. 結果を返す
    
    注意:
        - プレビューモードでは、実際の置換は行われません
        - 置換実行時は、自動的にバックアップファイル（.bak）が作成されます
        - Excelファイルとテキストファイルの両方に対応しています
    """
    try:
        data = request.json
        folder_path = data.get('folder_path', '')
        search_pattern = data.get('search_pattern', '')
        replace_pattern = data.get('replace_pattern', '')
        use_regex = data.get('use_regex', False)
        file_extensions = data.get('file_extensions', ['.txt', '.csv', '.html', '.js', '.ts', '.tsx', '.jsx', '.py', '.json', '.xml', '.css'])
        preview_only = data.get('preview_only', True)  # プレビューのみか、実際に置換するか
        
        if not folder_path:
            return jsonify({'success': False, 'error': 'フォルダパスが指定されていません'}), 400
        
        if not search_pattern:
            return jsonify({'success': False, 'error': '検索パターンが指定されていません'}), 400
        
        folder = Path(folder_path)
        if not folder.exists() or not folder.is_dir():
            return jsonify({'success': False, 'error': '指定されたフォルダが見つかりません'}), 404
        
        # 対象ファイルを取得
        target_files = []
        for ext in file_extensions:
            target_files.extend(list(folder.glob(f'*{ext}')))
            target_files.extend(list(folder.rglob(f'**/*{ext}')))  # サブディレクトリも検索
        
        # 重複を除去
        target_files = list(set(target_files))
        
        if not target_files:
            return jsonify({'success': False, 'error': '対象ファイルが見つかりませんでした'}), 404
        
        results = []
        total_replacements = 0
        
        # 正規表現のコンパイル
        if use_regex:
            try:
                pattern = re.compile(search_pattern)
            except re.error as e:
                return jsonify({'success': False, 'error': f'正規表現エラー: {str(e)}'}), 400
        else:
            # 通常の文字列検索（エスケープ処理）
            escaped_pattern = re.escape(search_pattern)
            pattern = re.compile(escaped_pattern)
        
        # 各ファイルを処理
        for file_path in target_files:
            try:
                # Excelファイルかどうかを判定
                is_excel = file_path.suffix.lower() in ['.xlsx', '.xls']
                
                if is_excel:
                    # Excelファイルの処理
                    try:
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        file_result = {
                            'file_path': str(file_path),
                            'file_name': file_path.name,
                            'matches': [],
                            'total_matches': 0,
                            'replaced': False
                        }
                        
                        # バックアップを作成（置換実行前）
                        if not preview_only:
                            backup_path = file_path.with_suffix(file_path.suffix + '.bak')
                            shutil.copy2(file_path, backup_path)
                            file_result['backup_path'] = str(backup_path)
                        
                        # 各シートを処理
                        for sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            
                            # 各セルを走査
                            for row in ws.iter_rows():
                                for cell in row:
                                    if cell.value is None:
                                        continue
                                    
                                    # セルの値を文字列に変換
                                    cell_value = str(cell.value)
                                    
                                    # 検索実行
                                    matches = list(pattern.finditer(cell_value))
                                    
                                    if matches:
                                        for match in matches:
                                            file_result['total_matches'] += 1
                                            file_result['matches'].append({
                                                'line': cell.row,
                                                'start': match.start(),
                                                'end': match.end(),
                                                'match_text': match.group(),
                                                'line_content': cell_value,
                                                'context_before': cell_value[max(0, match.start()-50):match.start()],
                                                'context_after': cell_value[match.end():min(len(cell_value), match.end()+50)],
                                                'sheet': sheet_name,
                                                'column': cell.column_letter
                                            })
                                            
                                            # 置換実行（プレビューモードでない場合）
                                            if not preview_only:
                                                # セルの値を置換
                                                if use_regex:
                                                    new_value = pattern.sub(replace_pattern, cell_value)
                                                else:
                                                    new_value = cell_value.replace(search_pattern, replace_pattern)
                                                
                                                # セルに新しい値を設定
                                                cell.value = new_value
                                                total_replacements += 1
                        
                        # Excelファイルを保存（置換実行した場合）
                        if not preview_only and file_result['total_matches'] > 0:
                            wb.save(file_path)
                            file_result['replaced'] = True
                        
                        # 結果が1つでもあれば追加
                        if file_result['total_matches'] > 0:
                            results.append(file_result)
                        
                        wb.close()
                        
                    except Exception as excel_error:
                        results.append({
                            'file_path': str(file_path),
                            'file_name': file_path.name,
                            'error': f'Excelファイル処理エラー: {str(excel_error)}',
                            'matches': [],
                            'total_matches': 0
                        })
                else:
                    # テキストファイルの処理（既存の処理）
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                    
                    # 検索実行
                    matches = list(pattern.finditer(content))
                    
                    if matches:
                        file_result = {
                            'file_path': str(file_path),
                            'file_name': file_path.name,
                            'matches': [],
                            'total_matches': len(matches),
                            'replaced': False
                        }
                        
                        # 各マッチの情報を取得
                        for match in matches:
                            start_pos = match.start()
                            end_pos = match.end()
                            
                            # 該当行を取得
                            line_number = content[:start_pos].count('\n') + 1
                            line_start = content.rfind('\n', 0, start_pos) + 1
                            line_end = content.find('\n', end_pos)
                            if line_end == -1:
                                line_end = len(content)
                            line_content = content[line_start:line_end]
                            
                            file_result['matches'].append({
                                'line': line_number,
                                'start': start_pos,
                                'end': end_pos,
                                'match_text': match.group(),
                                'line_content': line_content,
                                'context_before': content[max(0, start_pos-50):start_pos],
                                'context_after': content[end_pos:min(len(content), end_pos+50)]
                            })
                        
                        # 置換実行（プレビューモードでない場合）
                        if not preview_only:
                            # バックアップを作成
                            backup_path = file_path.with_suffix(file_path.suffix + '.bak')
                            shutil.copy2(file_path, backup_path)
                            
                            # 置換実行
                            if use_regex:
                                new_content = pattern.sub(replace_pattern, content)
                            else:
                                new_content = content.replace(search_pattern, replace_pattern)
                            
                            # ファイルに書き込み
                            with open(file_path, 'w', encoding='utf-8') as f:
                                f.write(new_content)
                            
                            file_result['replaced'] = True
                            file_result['backup_path'] = str(backup_path)
                            total_replacements += len(matches)
                        
                        results.append(file_result)
                    
            except Exception as e:
                results.append({
                    'file_path': str(file_path),
                    'file_name': file_path.name,
                    'error': str(e),
                    'matches': [],
                    'total_matches': 0
                })
        
        return jsonify({
            'success': True,
            'results': results,
            'total_files': len(target_files),
            'files_with_matches': len([r for r in results if r.get('total_matches', 0) > 0]),
            'total_replacements': total_replacements,
            'preview_only': preview_only
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-file-path', methods=['POST'])
def get_file_path():
    """
    アップロードされたファイルから絶対パスを取得するAPIエンドポイント
    
    注意: ブラウザのセキュリティ制限により、元のファイルパスは取得できません。
    このエンドポイントは一時ファイルのパスを返しますが、実際の用途では
    バックエンドのフォルダ選択ダイアログ（/api/browse-folder）を使用することを推奨します。
    
    リクエスト:
        POST /api/get-file-path
        Content-Type: multipart/form-data
        Form Data:
            - file: アップロードされたファイル
    
    レスポンス:
        {
            "success": true,
            "file_path": "一時ファイルのパス",
            "parent_dir": "親ディレクトリのパス",
            "filename": "ファイル名",
            "message": "注意メッセージ"
        }
    
    注意:
        - 返されるパスは一時ファイルのパスであり、元のファイルのパスではありません
        - ブラウザのセキュリティ制限により、元のファイルパスは取得できません
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'ファイルが指定されていません'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'ファイルが指定されていません'}), 400
        
        # 一時ファイルとして保存してパスを取得
        import tempfile
        original_filename = file.filename
        temp_dir = tempfile.gettempdir()
        temp_file = os.path.join(temp_dir, file.filename)
        file.save(temp_file)
        
        # 絶対パスを取得
        abs_path = os.path.abspath(temp_file)
        
        # ファイルの親ディレクトリのパスを返す
        parent_dir = os.path.dirname(abs_path)
        
        # 一時ファイルを削除
        try:
            os.remove(temp_file)
        except:
            pass
        
        return jsonify({
            'success': True,
            'file_path': abs_path,
            'parent_dir': parent_dir,
            'filename': original_filename,
            'message': '注意: これは一時ファイルのパスです。元のファイルのパスを取得するには、ファイルを直接ドロップするか、フォルダパスを手動で入力してください。'
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error in get_file_path: {error_trace}")
        return jsonify({
            'success': False,
            'error': f'ファイルパス取得中にエラーが発生しました: {str(e)}'
        }), 500


@app.route('/api/browse-folder', methods=['POST'])
def browse_folder():
    """
    フォルダ選択ダイアログを開くAPIエンドポイント（Windows環境）
    
    Windows環境で、GUIフォルダ選択ダイアログを開いてフォルダを選択します。
    選択されたフォルダの完全パスを返します。
    
    注意:
        - Webサーバー環境（Vercelなど）ではGUIダイアログが開けないため、
          この機能は制限されます
        - ローカル環境（Windows）でのみ動作します
        - 環境変数DEFAULT_SEARCH_FOLDERが設定されている場合は、そのフォルダを返します
    
    リクエスト:
        POST /api/browse-folder
    
    レスポンス:
        成功時 (200):
        {
            "success": true,
            "folder_path": "選択されたフォルダのパス",
            "message": "フォルダが選択されました: ..."
        }
        
        エラー時 (500):
        {
            "success": false,
            "error": "エラーメッセージ"
        }
    
    処理の流れ:
        1. 環境変数DEFAULT_SEARCH_FOLDERを確認
        2. Windows環境の場合:
           - Tkinterを使用してフォルダ選択ダイアログを開く
           - 選択されたフォルダのパスを返す
        3. その他の環境:
           - エラーメッセージを返す
    
    制限事項:
        - GUI環境が利用できない場合（サーバー環境など）は動作しません
        - Windows環境以外では動作しません
    """
    try:
        # Vercel環境やサーバー環境ではGUIダイアログを開くことができない
        # 適切なエラーメッセージを返す
        if os.environ.get('VERCEL'):
            response = jsonify({
                'success': False,
                'error': 'Vercel環境ではフォルダ選択ダイアログは利用できません。フォルダパスを手動で入力してください。',
                'suggestion': 'フォルダパス入力欄に直接パスを入力するか、Excelファイルをドラッグ&ドロップしてください。'
            })
            response.headers['Content-Type'] = 'application/json'
            return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示
        
        # 環境変数からデフォルトフォルダを取得（設定されている場合）
        default_folder = os.environ.get('DEFAULT_SEARCH_FOLDER', '')
        
        if default_folder and Path(default_folder).exists():
            response = jsonify({
                'success': True,
                'folder_path': default_folder,
                'message': 'デフォルトフォルダを使用します'
            })
            response.headers['Content-Type'] = 'application/json'
            return response
        
        # GUIダイアログを試みる（ローカル環境でのみ動作）
        try:
            import tkinter as tk
            from tkinter import filedialog
            
            # ディスプレイが利用可能かチェック
            if platform.system() == 'Windows':
                try:
                    # Tkinterのルートウィンドウを非表示で作成
                    root = tk.Tk()
                    root.withdraw()  # メインウィンドウを非表示
                    root.attributes('-topmost', True)  # 最前面に表示
                    
                    # フォルダ選択ダイアログを開く
                    folder_path = filedialog.askdirectory(title='検索対象フォルダを選択')
                    
                    root.destroy()
                    
                    if folder_path:
                        # 完全パスを正規化
                        folder_path = os.path.abspath(folder_path)
                        response = jsonify({
                            'success': True,
                            'folder_path': folder_path,
                            'message': f'フォルダが選択されました: {folder_path}'
                        })
                        response.headers['Content-Type'] = 'application/json'
                        return response
                    else:
                        response = jsonify({
                            'success': False,
                            'error': 'フォルダが選択されませんでした'
                        })
                        response.headers['Content-Type'] = 'application/json'
                        return response
                except Exception as tk_error:
                    # Tkinter関連のエラー
                    app.logger.error(f"Tkinter error: {str(tk_error)}")
                    response = jsonify({
                        'success': False,
                        'error': 'GUI環境が利用できません。フォルダパスを手動で入力してください。',
                        'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
                    })
                    response.headers['Content-Type'] = 'application/json'
                    return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示
            else:
                response = jsonify({
                    'success': False,
                    'error': 'フォルダ選択機能はWindows環境でのみ利用可能です',
                    'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
                })
                response.headers['Content-Type'] = 'application/json'
                return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示
                
        except ImportError:
            # tkinterが利用できない場合
            response = jsonify({
                'success': False,
                'error': 'フォルダ選択機能は利用できません。フォルダパスを手動で入力してください。',
                'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
            })
            response.headers['Content-Type'] = 'application/json'
            return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示
        except Exception as e:
            # GUI関連のエラー
            error_msg = str(e)
            app.logger.error(f"Browse folder error: {error_msg}")
            if 'display' in error_msg.lower() or 'DISPLAY' in error_msg:
                response = jsonify({
                    'success': False,
                    'error': 'GUI環境が利用できません。フォルダパスを手動で入力してください。',
                    'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
                })
            else:
                response = jsonify({
                    'success': False,
                    'error': f'フォルダ選択中にエラーが発生しました: {error_msg}',
                    'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
                })
            response.headers['Content-Type'] = 'application/json'
            return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Unexpected error in browse_folder: {error_trace}")
        response = jsonify({
            'success': False,
            'error': f'予期しないエラーが発生しました: {str(e)}',
            'suggestion': 'フォルダパス入力欄に直接パスを入力してください。'
        })
        response.headers['Content-Type'] = 'application/json'
        return response, 200  # 200を返して、フロントエンドでエラーメッセージを表示


@app.route('/api/get-folder-path', methods=['POST'])
def get_folder_path():
    """
    アップロードされたファイルからフォルダの完全パスを取得するAPIエンドポイント
    
    注意: ブラウザのセキュリティ制限により、元のファイルパスは取得できません。
    このエンドポイントは、バックエンドのフォルダ選択ダイアログ（/api/browse-folder）
    を使用することを推奨します。
    
    リクエスト:
        POST /api/get-folder-path
        Content-Type: multipart/form-data
        Form Data:
            - file: アップロードされたファイル
            - folder_name: フォルダ名（オプション）
    
    レスポンス:
        {
            "success": false,
            "error": "ブラウザのセキュリティ制限により、フォルダの完全パスを取得できません。...",
            "suggestion": "バックエンドのフォルダ選択ダイアログ（/api/browse-folder）を使用してください"
        }
    
    注意:
        - このエンドポイントは常にエラーを返します
        - ブラウザのセキュリティ制限により、元のファイルパスは取得できません
        - /api/browse-folderエンドポイントの使用を推奨します
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'ファイルが指定されていません'}), 400
        
        file = request.files['file']
        folder_name = request.form.get('folder_name', '')
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'ファイルが指定されていません'}), 400
        
        # ブラウザのセキュリティ制限により、元のファイルパスは取得できません
        # 代わりに、バックエンドのフォルダ選択ダイアログを使用することを推奨
        return jsonify({
            'success': False,
            'error': 'ブラウザのセキュリティ制限により、フォルダの完全パスを取得できません。\n「フォルダ選択」ボタンを使用して、サーバー側でフォルダを選択してください。',
            'folder_name': folder_name,
            'suggestion': 'バックエンドのフォルダ選択ダイアログ（/api/browse-folder）を使用してください'
        })
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error in get_folder_path: {error_trace}")
        return jsonify({
            'success': False,
            'error': f'フォルダパス取得中にエラーが発生しました: {str(e)}'
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """
    ヘルスチェックAPIエンドポイント
    
    サーバーが正常に動作しているかを確認するために使用します。
    デプロイ環境やロードバランサーからのヘルスチェックに使用できます。
    
    リクエスト:
        GET /api/health
    
    レスポンス:
        {
            "status": "ok",
            "message": "Excel Search API is running"
        }
    
    用途:
        - サーバーの稼働状況確認
        - デプロイ後の動作確認
        - ロードバランサーや監視ツールからのヘルスチェック
    """
    return jsonify({'status': 'ok', 'message': 'Excel Search API is running'})


# ============================================================================
# メイン実行部分
# ============================================================================

if __name__ == '__main__':
    """
    スクリプトが直接実行された場合の処理
    
    この部分は、app.pyが直接実行された場合（python app.py）にのみ実行されます。
    VercelなどのServerless環境では実行されません。
    """
    import logging
    
    # ロギングの設定
    # INFOレベル以上のログを出力
    logging.basicConfig(level=logging.INFO)
    app.logger.setLevel(logging.INFO)
    
    # Vercel環境ではサーバーを起動しない
    # Vercel環境では、api/index.pyのハンドラーがリクエストを処理するため
    if not os.environ.get('VERCEL'):
        # 環境変数から設定を取得（デフォルト値あり）
        # これにより、環境に応じて設定を変更できる
        port = int(os.environ.get('FLASK_PORT', '5001'))  # ポート番号（デフォルト: 5001）
        debug_mode = os.environ.get('FLASK_DEBUG', 'True').lower() in ('true', '1', 'yes')  # デバッグモード
        host = os.environ.get('FLASK_HOST', '0.0.0.0')  # ホストアドレス（デフォルト: 0.0.0.0 = すべてのインターフェース）
        
        # 起動情報を表示
        print(f"Starting Flask server on http://{host}:{port}")
        print(f"Debug mode: {debug_mode}")
        print("API endpoints:")
        print("  - POST /api/search")  # フォルダ内のExcelファイルを検索
        print("  - POST /api/search-files")  # アップロードされたExcelファイルを検索
        print("  - POST /api/get-cell-details")  # セルの詳細情報を取得
        print("  - POST /api/open-excel-file")  # Excelファイルを開く
        print("  - POST /api/search-replace")  # 一括検索・置換
        print("  - GET /api/health")  # ヘルスチェック
        print("\n環境変数で設定を変更できます:")
        print("  - FLASK_PORT: ポート番号（デフォルト: 5001）")
        print("  - FLASK_DEBUG: デバッグモード（デフォルト: True）")
        print("  - FLASK_HOST: ホスト（デフォルト: 0.0.0.0）")
        print("  - DEFAULT_SEARCH_FOLDER: デフォルト検索フォルダ（オプション）")
        
        # Flaskサーバーを起動
        # debug=debug_mode: デバッグモードの設定（Trueの場合、コード変更時に自動リロード）
        # port=port: ポート番号
        # host=host: ホストアドレス（0.0.0.0で全てのインターフェースからアクセス可能）
        app.run(debug=debug_mode, port=port, host=host)
